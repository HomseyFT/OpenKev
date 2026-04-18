"""XLSX import/export via openpyxl.

Import policy:

* Cell sources come in as-is — formulas keep their leading ``=`` and are
  re-evaluated by Kevcel's engine. Excel's cached evaluated values are
  intentionally discarded.
* Styles are translated on a best-effort basis: bold, italic, underline, font
  family/size, font color, fill color, horizontal alignment.

Export policy mirrors the above. Unsupported features (borders, merged cells)
are simply skipped — see the module docstring in :mod:`kevcel.core.workbook`
for the v1 scope.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook as XLWorkbook, load_workbook as xl_load
from openpyxl.styles import Alignment, Font, PatternFill

from apps.Kevcel.core.styles import CellStyle, HAlign
from apps.Kevcel.core.workbook import Cell, Sheet, Workbook


# ---- Import ---------------------------------------------------------------


def import_xlsx(path: str | Path) -> Workbook:
    xl = xl_load(filename=str(path), data_only=False)
    sheet_names = xl.sheetnames or ["Sheet1"]
    wb = Workbook(sheet_names)
    for s_idx, name in enumerate(sheet_names):
        xl_sheet = xl[name]
        sheet = wb.sheets[s_idx]
        for row in xl_sheet.iter_rows():
            for xl_cell in row:
                raw = xl_cell.value
                if raw is None:
                    continue
                source = _xl_cell_source(raw)
                style = _xl_style_to_kev(xl_cell)
                r, c = xl_cell.row - 1, xl_cell.column - 1
                sheet.cells[(r, c)] = Cell(source=source, style=style)
                sheet.ensure_bounds(r, c)
    wb.recalculate_all()
    return wb


def _xl_cell_source(raw: object) -> str:
    # openpyxl returns formulas as strings starting with '='
    if isinstance(raw, str):
        return raw
    if isinstance(raw, bool):
        return "TRUE" if raw else "FALSE"
    return str(raw)


def _xl_style_to_kev(xl_cell) -> CellStyle:
    style = CellStyle()
    font = xl_cell.font
    if font is not None:
        if font.bold:
            style = style.with_bold(True)
        if font.italic:
            style = style.with_italic(True)
        if font.underline and font.underline != "none":
            style = style.with_underline(True)
        if font.name:
            style = style.with_font(family=font.name)
        if font.size:
            try:
                style = style.with_font(size=int(font.size))
            except (TypeError, ValueError):
                pass
        if font.color and font.color.type == "rgb" and font.color.rgb:
            style = style.with_colors(font_color=_hex_from_argb(font.color.rgb))
    fill = xl_cell.fill
    if fill is not None and isinstance(fill, PatternFill) and fill.fgColor is not None:
        if fill.fgColor.type == "rgb" and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
            style = style.with_colors(fill_color=_hex_from_argb(fill.fgColor.rgb))
    alignment = xl_cell.alignment
    if alignment is not None and alignment.horizontal:
        mapping = {"left": HAlign.LEFT, "center": HAlign.CENTER, "right": HAlign.RIGHT}
        if alignment.horizontal in mapping:
            style = style.with_alignment(h=mapping[alignment.horizontal])
    return style


def _hex_from_argb(argb: str) -> str:
    """Convert an openpyxl AARRGGBB color string to CSS ``#RRGGBB``."""
    if len(argb) == 8:
        argb = argb[2:]
    return "#" + argb.lower()


# ---- Export ---------------------------------------------------------------


def export_xlsx(wb: Workbook, path: str | Path) -> None:
    xl = XLWorkbook()
    # Remove openpyxl's default sheet; we'll recreate ours.
    default = xl.active
    xl.remove(default)
    for sheet in wb.sheets:
        xl_sheet = xl.create_sheet(title=sheet.name)
        _write_sheet(sheet, xl_sheet)
    xl.save(str(path))


def _write_sheet(sheet: Sheet, xl_sheet) -> None:
    for (r, c), cell in sheet.cells.items():
        xl_cell = xl_sheet.cell(row=r + 1, column=c + 1)
        xl_cell.value = _kev_source_to_xl(cell.source)
        _kev_style_to_xl(cell.style, xl_cell)


def _kev_source_to_xl(source: str) -> object:
    # Formulas pass through. Everything else is sent as plain text; Excel will
    # happily coerce numeric-looking strings on first open.
    return source


def _kev_style_to_xl(style: CellStyle, xl_cell) -> None:
    font_kwargs = {}
    if style.bold:
        font_kwargs["bold"] = True
    if style.italic:
        font_kwargs["italic"] = True
    if style.underline:
        font_kwargs["underline"] = "single"
    if style.font_family:
        font_kwargs["name"] = style.font_family
    if style.font_size:
        font_kwargs["size"] = style.font_size
    if style.font_color:
        font_kwargs["color"] = _css_to_argb(style.font_color)
    if font_kwargs:
        xl_cell.font = Font(**font_kwargs)
    if style.fill_color:
        xl_cell.fill = PatternFill(
            patternType="solid",
            fgColor=_css_to_argb(style.fill_color),
            bgColor=_css_to_argb(style.fill_color),
        )
    if style.h_align in (HAlign.LEFT, HAlign.CENTER, HAlign.RIGHT):
        xl_cell.alignment = Alignment(horizontal=style.h_align.value)


def _css_to_argb(css: str) -> str:
    """Convert a ``#RRGGBB`` CSS color to ``AARRGGBB`` for openpyxl."""
    hex_part = css.lstrip("#")
    if len(hex_part) == 6:
        return ("FF" + hex_part).upper()
    return hex_part.upper()
